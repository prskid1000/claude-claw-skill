"""claw xlsx read — dump cells / ranges as JSON / CSV / TSV."""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

import click

from claw.common import EXIT_INPUT, RangeSelector, die


@click.command(name="read")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--sheet", default=None, help="Sheet name or 1-based index; default first.")
@click.option("--range", "a1_range", default=None, help="A1 range, e.g. A1:D10.")
@click.option("--values-only", is_flag=True, help="Return just a 2D array, no coords.")
@click.option("--formulas", is_flag=True, help="Emit formula strings instead of values.")
@click.option("--json", "as_json", is_flag=True, help="Emit JSON on stdout.")
@click.option("--csv", "as_csv", is_flag=True, help="Emit CSV on stdout.")
@click.option("--tsv", "as_tsv", is_flag=True, help="Emit TSV on stdout.")
@click.option("--stream", is_flag=True, help="Use read_only for large files.")
def read(src: Path, sheet: str | None, a1_range: str | None,
         values_only: bool, formulas: bool,
         as_json: bool, as_csv: bool, as_tsv: bool, stream: bool) -> None:
    """Print cell values as JSON (default), CSV, or TSV."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    wb = load_workbook(src, read_only=stream, data_only=not formulas)
    if sheet is None:
        ws = wb.worksheets[0]
    elif sheet.isdigit():
        ws = wb.worksheets[int(sheet) - 1]
    else:
        ws = wb[sheet]

    if a1_range:
        r1, c1, r2, c2 = RangeSelector(a1_range).resolve()
        cells = ws.iter_rows(min_row=r1, min_col=c1,
                             max_row=r2, max_col=c2, values_only=True)
    else:
        cells = ws.iter_rows(values_only=True)
    rows = [list(r) for r in cells]

    if as_csv or as_tsv:
        delim = "\t" if as_tsv else ","
        w = csv.writer(sys.stdout, delimiter=delim, lineterminator="\n")
        for row in rows:
            w.writerow(["" if v is None else v for v in row])
        return

    if values_only or as_json or not (as_csv or as_tsv):
        sys.stdout.write(json.dumps(rows, default=str, ensure_ascii=False) + "\n")
