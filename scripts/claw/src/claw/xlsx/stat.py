"""claw xlsx stat — per-column summary statistics."""

from __future__ import annotations

import math
import statistics
from pathlib import Path

import click

from claw.common import EXIT_INPUT, RangeSelector, die, emit_json


@click.command(name="stat")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--sheet", required=True)
@click.option("--range", "a1_range", default=None)
@click.option("--columns", default=None, help="Comma-separated header names.")
@click.option("--json", "as_json", is_flag=True)
def stat(src: Path, sheet: str, a1_range: str | None,
         columns: str | None, as_json: bool) -> None:
    """Report min / max / mean / stddev / distinct / null counts per column."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb[sheet]

    if a1_range:
        r1, c1, r2, c2 = RangeSelector(a1_range).resolve()
        rows = list(ws.iter_rows(min_row=r1, min_col=c1,
                                 max_row=r2, max_col=c2, values_only=True))
    else:
        rows = list(ws.iter_rows(values_only=True))

    if not rows:
        die("no rows to summarize", code=EXIT_INPUT, as_json=as_json)
    headers = [str(h) if h is not None else f"col{i+1}" for i, h in enumerate(rows[0])]
    data = rows[1:]
    wanted = set(headers)
    if columns:
        wanted = {c.strip() for c in columns.split(",")}

    summary = []
    for i, h in enumerate(headers):
        if h not in wanted:
            continue
        col = [r[i] for r in data]
        nulls = sum(1 for v in col if v is None or v == "")
        nonnull = [v for v in col if v is not None and v != ""]
        distinct = len({str(v) for v in nonnull})
        nums = [v for v in nonnull if isinstance(v, (int, float))
                and not (isinstance(v, float) and math.isnan(v))]
        entry = {"column": h, "count": len(col), "nulls": nulls, "distinct": distinct}
        if nums:
            entry.update({
                "min": min(nums), "max": max(nums),
                "mean": round(statistics.fmean(nums), 6),
                "stddev": round(statistics.pstdev(nums), 6) if len(nums) > 1 else 0,
            })
        summary.append(entry)

    if as_json:
        emit_json(summary)
    else:
        for e in summary:
            click.echo(e)
