"""claw xlsx meta — get/set core workbook properties."""

from __future__ import annotations

from pathlib import Path

import click

from claw.common import EXIT_INPUT, common_output_options, die, emit_json, safe_write


@click.group(name="meta")
def meta() -> None:
    """Read or write core workbook properties."""


@meta.command(name="get")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--json", "as_json", is_flag=True)
def meta_get(src: Path, as_json: bool) -> None:
    """Print core properties."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    wb = load_workbook(src, read_only=True)
    p = wb.properties
    info = {
        "title": p.title, "creator": p.creator, "subject": p.subject,
        "keywords": p.keywords, "description": p.description,
        "company": getattr(p, "company", None),
        "lastModifiedBy": p.lastModifiedBy,
        "created": p.created, "modified": p.modified,
    }
    if as_json:
        emit_json(info)
    else:
        for k, v in info.items():
            click.echo(f"{k}: {v}")


@meta.command(name="set")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--title", default=None)
@click.option("--author", default=None, help="Sets 'creator'.")
@click.option("--subject", default=None)
@click.option("--keywords", default=None)
@click.option("--description", default=None)
@click.option("--company", default=None)
@common_output_options
def meta_set(src: Path, title: str | None, author: str | None, subject: str | None,
             keywords: str | None, description: str | None, company: str | None,
             force: bool, backup: bool, as_json: bool, dry_run: bool,
             quiet: bool, verbose: bool, mkdir: bool) -> None:
    """Set core properties (title/author/subject/etc)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    updates = {"title": title, "creator": author, "subject": subject,
               "keywords": keywords, "description": description}

    if dry_run:
        click.echo(f"would set: {updates}")
        return

    wb = load_workbook(src)
    for k, v in updates.items():
        if v is not None:
            setattr(wb.properties, k, v)
    if company is not None and hasattr(wb.properties, "company"):
        wb.properties.company = company

    def _save(f):
        wb.save(f)

    safe_write(src, _save, force=True, backup=backup, mkdir=mkdir)

    if as_json:
        emit_json({"path": str(src), "updates": {k: v for k, v in updates.items() if v is not None}})
    elif not quiet:
        click.echo(f"updated metadata on {src}")
