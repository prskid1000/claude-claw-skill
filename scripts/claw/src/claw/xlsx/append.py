"""claw xlsx append — append rows from CSV / JSON to a sheet."""

from __future__ import annotations

from pathlib import Path

import click

from claw.common import (
    EXIT_INPUT, common_output_options, die, emit_json, read_rows, safe_write,
)


@click.command(name="append")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--sheet", required=True, help="Target sheet (created if missing).")
@click.option("--data", "data_src", required=True, help="CSV/JSON file or - for stdin.")
@click.option("--types", "types_mode", type=click.Choice(["infer", "text"]), default="infer")
@click.option("--stream", is_flag=True, help="No-op placeholder (openpyxl append is in-memory).")
@common_output_options
def append(src: Path, sheet: str, data_src: str, types_mode: str, stream: bool,
           force: bool, backup: bool, as_json: bool, dry_run: bool,
           quiet: bool, verbose: bool, mkdir: bool) -> None:
    """Append rows from a CSV/JSON file (or stdin) to a sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    rows = read_rows(data_src, header=True)
    if dry_run:
        click.echo(f"would append {len(rows)} rows to {sheet} in {src}")
        return

    wb = load_workbook(src)
    if sheet not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet)
        if rows and isinstance(rows[0], dict):
            ws.append(list(rows[0].keys()))
    else:
        ws = wb[sheet]

    for r in rows:
        if isinstance(r, dict):
            values = list(r.values())
        else:
            values = list(r)
        if types_mode == "text":
            values = ["" if v is None else str(v) for v in values]
        ws.append(values)

    def _save(f):
        wb.save(f)

    safe_write(src, _save, force=True, backup=backup, mkdir=mkdir)

    if as_json:
        emit_json({"path": str(src), "sheet": sheet, "rows_appended": len(rows)})
    elif not quiet:
        click.echo(f"appended {len(rows)} rows to {sheet}")
