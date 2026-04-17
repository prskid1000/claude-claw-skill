"""claw xlsx to-csv — export a single sheet to CSV."""

from __future__ import annotations

import csv
import sys
from pathlib import Path

import click

from claw.common import (
    EXIT_INPUT, RangeSelector, common_output_options, die, emit_json, safe_write,
)


@click.command(name="to-csv")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--sheet", required=True, help="Sheet name or 1-based index.")
@click.option("--out", "dst", default="-", help="Output CSV (or - for stdout).")
@click.option("--delimiter", default=",")
@click.option("--range", "a1_range", default=None)
@click.option("--encoding", default="utf-8")
@common_output_options
def to_csv(src: Path, sheet: str, dst: str, delimiter: str, a1_range: str | None,
           encoding: str,
           force: bool, backup: bool, as_json: bool, dry_run: bool,
           quiet: bool, verbose: bool, mkdir: bool) -> None:
    """Export a sheet to CSV (file or stdout)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb.worksheets[int(sheet) - 1] if sheet.isdigit() else wb[sheet]

    if a1_range:
        r1, c1, r2, c2 = RangeSelector(a1_range).resolve()
        iterator = ws.iter_rows(min_row=r1, min_col=c1,
                                max_row=r2, max_col=c2, values_only=True)
    else:
        iterator = ws.iter_rows(values_only=True)

    rows = [["" if v is None else v for v in r] for r in iterator]

    if dry_run:
        click.echo(f"would export {len(rows)} rows → {dst}")
        return

    if dst == "-":
        w = csv.writer(sys.stdout, delimiter=delimiter, lineterminator="\n")
        for row in rows:
            w.writerow(row)
    else:
        def _writer(f):
            import io
            text = io.StringIO()
            w = csv.writer(text, delimiter=delimiter, lineterminator="\n")
            for row in rows:
                w.writerow(row)
            f.write(text.getvalue().encode(encoding))

        safe_write(Path(dst), _writer, force=force, backup=backup, mkdir=mkdir)

    if as_json:
        emit_json({"path": dst, "rows": len(rows)})
    elif not quiet and dst != "-":
        click.echo(f"wrote {dst} ({len(rows)} rows)")
