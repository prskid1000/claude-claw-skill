"""claw xlsx freeze — freeze top N rows / left N cols."""

from __future__ import annotations

from pathlib import Path

import click

from claw.common import EXIT_INPUT, common_output_options, die, emit_json, safe_write


@click.command(name="freeze")
@click.argument("src", type=click.Path(exists=True, path_type=Path))
@click.option("--sheet", required=True)
@click.option("--rows", default=0, type=int)
@click.option("--cols", default=0, type=int)
@click.option("--at", "at_cell", default=None, help="Direct freeze anchor, e.g. B2.")
@common_output_options
def freeze(src: Path, sheet: str, rows: int, cols: int, at_cell: str | None,
           force: bool, backup: bool, as_json: bool, dry_run: bool,
           quiet: bool, verbose: bool, mkdir: bool) -> None:
    """Freeze the top N rows and/or left N columns (or pass --at CELL)."""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        die("openpyxl not installed", code=EXIT_INPUT,
            hint="uv tool install 'claw[xlsx]'", as_json=as_json)

    if at_cell:
        anchor = at_cell
    else:
        anchor = f"{get_column_letter(cols + 1)}{rows + 1}"

    if dry_run:
        click.echo(f"would set freeze_panes={anchor} on {sheet}")
        return

    wb = load_workbook(src)
    ws = wb[sheet]
    ws.freeze_panes = anchor

    def _save(f):
        wb.save(f)

    safe_write(src, _save, force=True, backup=backup, mkdir=mkdir)

    if as_json:
        emit_json({"path": str(src), "sheet": sheet, "freeze_panes": anchor})
    elif not quiet:
        click.echo(f"froze panes at {anchor} on {sheet}")
