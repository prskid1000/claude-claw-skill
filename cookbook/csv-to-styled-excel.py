#!/usr/bin/env python3
"""
Cortex Example: csv-to-styled-excel
Description: Convert CSV to styled Excel with auto-filters and formatting
Tags: csv,excel,openpyxl,conversion
Captured: 2026-03-25
Source: csv-to-styled-excel.py

Usage:
  python ~/.claude/skills/cortex/cookbook/csv-to-styled-excel.py
"""

#!/usr/bin/env python3
"""
Convert a CSV file to a styled Excel workbook with auto-filters, column widths, and header formatting.

Usage:
    python csv-to-styled-excel.py input.csv [output.xlsx]
"""

import sys
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def csv_to_styled_excel(csv_path, xlsx_path=None):
    csv_path = Path(csv_path)
    if not xlsx_path:
        xlsx_path = csv_path.with_suffix(".xlsx")

    # Read CSV
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        print("Empty CSV file.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = csv_path.stem[:31]

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # Write data
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            else:
                if r_idx % 2 == 0:
                    cell.fill = alt_fill

    # Auto-filter
    if len(rows) > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{len(rows)}"

    # Auto-width
    for col_idx in range(1, len(rows[0]) + 1):
        max_len = 0
        for row in rows:
            if col_idx <= len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(str(xlsx_path))
    print(f"Saved: {xlsx_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python csv-to-styled-excel.py input.csv [output.xlsx]")
        sys.exit(1)
    csv_to_styled_excel(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
